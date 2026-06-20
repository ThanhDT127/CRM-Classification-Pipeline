"""
CRM Classification Pipeline - Step 4: Merge LLM Fills
======================================================
Merge llm_fills.json results back into CRM_classified.xlsx
to produce CRM_classified_with_LLM.xlsx.
"""

import json
import pandas as pd
from pathlib import Path
from typing import Any, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    PATH_OUTPUT_REGEX,
    PATH_OUTPUT_LLM,
    OUT_LLM_JSON,
    OUTPUT_COLUMNS,
    LLM_TARGET_COLS,
    COL_CURRENT_STATUS,
    INPUT_TEXT_COLUMNS,
)


def apply_excel_styling(file_path: Path):
    """
    Apply premium styling to output columns in the Excel spreadsheet.
    Creates a double-row header structure:
    - Row 1: Merged major group name (e.g. AETT, Khách Hàng, Hoạt Động CRM)
    - Row 2: Minor column name
    - Non-group columns: Spanned vertically across Row 1 and Row 2.
    """
    print(f"Applying premium two-row header styling to {file_path.name}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Color mapping matching user's reference image
        group_colors = {
            "Hoạt Động CRM": "FFF2CC",          # Yellow / Gold
            "AETT": "E2EFDA",                   # Light Green
            "Khách Hàng": "E8D5F5",              # Light Purple
            "Kế Hoạch": "FCE4CC",               # Light Orange / Peach
            "Đối Thủ Cạnh Tranh": "D6EAF8"      # Light Blue
        }

        # Lighter data-cell fills for classification columns
        group_data_colors = {
            "Hoạt Động CRM": "FFFBEA",
            "AETT": "F2F9EE",
            "Khách Hàng": "F5EEFA",
            "Kế Hoạch": "FEF3EA",
            "Đối Thủ Cạnh Tranh": "EBF5FB"
        }

        # Define standard fonts, alignments, and borders
        font_major = Font(name="Segoe UI", size=11, bold=True)
        font_minor = Font(name="Segoe UI", size=10, bold=True)
        font_data = Font(name="Segoe UI", size=10)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Insert new row 1
        ws.insert_rows(1, 1)
        
        # Store column info: col_idx -> (major, minor)
        col_info = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(row=2, column=col_idx).value or "")
            if cell_val.startswith("[") and "]" in cell_val:
                parts = cell_val.split("] ", 1)
                major = parts[0][1:] # strip the "["
                minor = parts[1]
                col_info[col_idx] = (major, minor)
                
                # Write to cells
                ws.cell(row=1, column=col_idx, value=major)
                ws.cell(row=2, column=col_idx, value=minor)
            else:
                col_info[col_idx] = (None, None)

        # Style header cells in row 1 and row 2
        for col_idx in range(1, ws.max_column + 1):
            major, minor = col_info.get(col_idx, (None, None))
            
            if major:
                fill_color = group_colors.get(major, "FFFFFF")
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                # Style row 1
                c1 = ws.cell(row=1, column=col_idx)
                c1.fill = fill
                c1.font = font_major
                c1.alignment = align_center
                c1.border = thin_border
                
                # Style row 2
                c2 = ws.cell(row=2, column=col_idx)
                c2.fill = fill
                c2.font = font_minor
                c2.alignment = align_center
                c2.border = thin_border
            else:
                # Non-group column: merge vertically
                val = ws.cell(row=2, column=col_idx).value
                ws.cell(row=1, column=col_idx, value=val)
                ws.cell(row=2, column=col_idx, value="")
                
                fill_neutral = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for r in (1, 2):
                    c = ws.cell(row=r, column=col_idx)
                    c.fill = fill_neutral
                    c.font = font_minor
                    c.alignment = align_center
                    c.border = thin_border
                    
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

            # Style data rows — apply light fill to classification columns
            if major:
                data_color = group_data_colors.get(major, "FFFFFF")
                data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
            else:
                data_fill = None

            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                if data_fill and major:
                    cell.fill = data_fill

        # Now perform the horizontal merges in row 1 for groups
        start_col = None
        current_major = None
        
        for col_idx in range(1, ws.max_column + 2):
            major = col_info.get(col_idx, (None, None))[0]
            if major != current_major:
                if current_major is not None and start_col is not None:
                    end_col = col_idx - 1
                    if end_col > start_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                if major is not None:
                    start_col = col_idx
                    current_major = major
                else:
                    start_col = None
                    current_major = None

        # Set row heights
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24

        # Auto-fit column widths (checking first 100 rows for performance)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                val_str = str(ws.cell(row=row_idx, column=col_idx).value or "")
                lines = val_str.split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(file_path)
        print(f"Styling and auto-fit applied successfully to {file_path.name}!")
    except Exception as e:
        print(f"Error applying styling to Excel: {e}")


def main():
    # 1. Load classified Excel
    print(f"Loading {PATH_OUTPUT_REGEX.name} ...")
    df = pd.read_excel(PATH_OUTPUT_REGEX)
    print(f"  Rows: {len(df)}")

    # 2. Load LLM fills
    if not OUT_LLM_JSON.exists():
        print(f"[ERROR] LLM fills not found: {OUT_LLM_JSON}. Run step3 first.")
        return

    fills_list = json.loads(OUT_LLM_JSON.read_text(encoding='utf-8'))
    print(f"  LLM fill items: {len(fills_list)}")

    # Build lookup by row_idx
    fills_by_idx: Dict[int, Dict[str, Any]] = {}
    for item in fills_list:
        rid = int(item.get('row_idx', 0))
        fills = item.get('fills', {})
        if isinstance(fills, dict):
            fills_by_idx[rid] = fills

    # 3. Ensure output columns exist and cast to object dtype to prevent Pandas dtype assignment errors
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None
        df[col] = df[col].astype(object)

    # 4. Merge fills
    merged_count = 0
    cleared_count = 0
    for idx, row in df.iterrows():
        rid = int(row.get('row_idx', 0))
        fills = fills_by_idx.get(rid)
        if not fills:
            continue
        for col in LLM_TARGET_COLS:
            if col not in fills:
                continue
            val = fills[col]
            cur = row.get(col)
            is_empty_or_mo_ho = pd.isna(cur) or str(cur).strip() == '' or str(cur).strip().lower() == 'mơ hồ'
            if is_empty_or_mo_ho:
                if val is not None:
                    s = str(val).strip()
                    if s:
                        df.at[idx, col] = s
                        merged_count += 1
                    else:
                        df.at[idx, col] = None
                        if str(cur).strip().lower() == 'mơ hồ':
                            cleared_count += 1
                else:
                    df.at[idx, col] = None
                    if str(cur).strip().lower() == 'mơ hồ':
                        cleared_count += 1

    # 5. Export
    df.to_excel(PATH_OUTPUT_LLM, index=False)
    print(f"\n[OK] Merge done!")
    print(f"  Total cells filled by LLM: {merged_count}")
    print(f"  Total 'mo ho' placeholders cleared: {cleared_count}")
    print(f"  Output: {PATH_OUTPUT_LLM.name}")

    # 5.1 Apply styling
    apply_excel_styling(PATH_OUTPUT_LLM)

    # 6. Summary stats
    print(f"\nFinal fill rates per classification column:")
    for col in OUTPUT_COLUMNS:
        if col not in df.columns:
            continue
        filled = df[col].notna().sum()
        filled -= (df[col].astype(str).str.strip() == '').sum()
        pct = filled / len(df) * 100
        col_safe = col.encode('ascii', errors='replace').decode('ascii')
        print(f"    {col_safe}: {filled}/{len(df)} ({pct:.1f}%)")


if __name__ == '__main__':
    main()
