import os
import sys
import json
import time
import shutil
import hashlib
from copy import copy

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import datetime
import traceback
import logging
from logging.handlers import RotatingFileHandler
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config
from classifier import (
    classify_row_regex,
    _compile_patterns,
    _clean
)
from llm import init_llm_client, call_llm_batch
from sharepoint import AuthProvider, SharePointClient
from notification import NotificationService

# --- SETUP LOGGING ---
config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
log_dir = config.PATH_OUTPUT / "logs"
log_dir.mkdir(parents=True, exist_ok=True)
log_file = log_dir / "automation.log"

logger = logging.getLogger("crm-automation")
logger.setLevel(logging.INFO)

CACHE_META_PROCESSED = "__DMS_PROCESSED__"
CACHE_META_CONTENT_HASH = "__DMS_CONTENT_HASH__"
CACHE_META_UPDATED_AT = "__DMS_UPDATED_AT__"
CACHE_META_COLUMNS = [
    CACHE_META_PROCESSED,
    CACHE_META_CONTENT_HASH,
    CACHE_META_UPDATED_AT,
]

# Create formatter
formatter = logging.Formatter("[%(asctime)s] %(levelname)s [%(name)s] %(message)s")

# Stdout handler
stdout_handler = logging.StreamHandler(sys.stdout)
stdout_handler.setFormatter(formatter)
logger.addHandler(stdout_handler)

# Rotating file handler (max 10MB per file, keep 5 files)
file_handler = RotatingFileHandler(log_file, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8")
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

def load_and_index_keywords():
    with open(config.PATH_KW_JSON, "r", encoding="utf-8") as f:
        kw = json.load(f)
        
    kw_index = {}
    competitor_brands = []
    
    for major, minors in (kw or {}).items():
        if not isinstance(minors, dict):
            continue
        for minor, subs in minors.items():
            if major == "Đối Thủ Cạnh Tranh" and minor == "Cạnh Tranh" and isinstance(subs, dict):
                brand_list = subs.get("Các Hãng cụ thể") or subs.get("Các Hãng đối thủ cạnh tranh")
                if isinstance(brand_list, list):
                    for b in brand_list:
                        for term in str(b).split(";"):
                            if term.strip():
                                competitor_brands.append(term.strip())
                continue
            if isinstance(subs, dict):
                col = config.col_name(major, minor)
                col_labels = {}
                for sub, kw_list in subs.items():
                    if not isinstance(sub, str) or not sub.strip():
                        continue
                    terms = []
                    if isinstance(kw_list, list):
                        for k in kw_list:
                            for t in str(k).split(";"):
                                if t.strip():
                                    terms.append(t.strip())
                    elif isinstance(kw_list, str):
                        terms = [t.strip() for t in kw_list.split(";") if t.strip()]
                    if terms:
                        col_labels[sub] = terms
                if col_labels:
                    kw_index[col] = col_labels
                    
    return kw_index, competitor_brands

def apply_excel_styling(file_path: Path):
    logger.info("Applying premium double-header styling to Excel...")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    group_colors = {
        "Hoạt Động CRM": "FFF2CC",
        "AETT": "E2EFDA",
        "Khách Hàng": "E8D5F5",
        "Kế Hoạch": "FCE4CC",
        "Đối Thủ Cạnh Tranh": "D6EAF8"
    }

    group_data_colors = {
        "Hoạt Động CRM": "FFFBEA",
        "AETT": "F2F9EE",
        "Khách Hàng": "F5EEFA",
        "Kế Hoạch": "FEF3EA",
        "Đối Thủ Cạnh Tranh": "EBF5FB"
    }

    font_major = Font(name="Segoe UI", size=11, bold=True)
    font_minor = Font(name="Segoe UI", size=10, bold=True)
    font_data = Font(name="Segoe UI", size=10)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    ws.insert_rows(1, 1)
    
    col_info = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_val = str(ws.cell(row=2, column=col_idx).value or "")
        if cell_val.startswith("[") and "]" in cell_val:
            parts = cell_val.split("] ", 1)
            major = parts[0][1:]
            minor = parts[1]
            col_info[col_idx] = (major, minor)
            ws.cell(row=1, column=col_idx, value=major)
            ws.cell(row=2, column=col_idx, value=minor)
        else:
            col_info[col_idx] = (None, None)

    for col_idx in range(1, ws.max_column + 1):
        major, minor = col_info.get(col_idx, (None, None))
        
        if major:
            fill_color = group_colors.get(major, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            c1 = ws.cell(row=1, column=col_idx)
            c1.fill = fill
            c1.font = font_major
            c1.alignment = align_center
            c1.border = thin_border
            
            c2 = ws.cell(row=2, column=col_idx)
            c2.fill = fill
            c2.font = font_minor
            c2.alignment = align_center
            c2.border = thin_border
        else:
            val = ws.cell(row=2, column=col_idx).value
            ws.cell(row=1, column=col_idx, value=val)
            ws.cell(row=2, column=col_idx, value="")
            
            fill_neutral = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for r in (1, 2):
                c = ws.cell(row=r, column=col_idx)
                c.fill = fill_neutral
                c.font = font_minor
                c.alignment = align_center
                c.border = thin_border
                
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

        if major:
            data_color = group_data_colors.get(major, "FFFFFF")
            data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
        else:
            data_fill = None

        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            if data_fill and major:
                cell.fill = data_fill

    start_col = None
    current_major = None
    for col_idx in range(1, ws.max_column + 2):
        major = col_info.get(col_idx, (None, None))[0]
        if major != current_major:
            if current_major is not None and start_col is not None:
                end_col = col_idx - 1
                if end_col > start_col:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            if major is not None:
                start_col = col_idx
                current_major = major
            else:
                start_col = None
                current_major = None

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            val_str = str(ws.cell(row=row_idx, column=col_idx).value or "")
            lines = val_str.split("\n")
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        
        # Check if this column is a classification output column
        header_val = str(ws.cell(row=2, column=col_idx).value or "")
        is_classification = any(header_val == c.split("] ")[-1] for c in config.OUTPUT_COLUMNS)
        if is_classification:
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)
        else:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 20)

    wb.save(file_path)
    logger.info("[OK] Excel styling successfully applied.")

def save_history_db_atomic(history_db: dict):
    import tempfile
    import os
    file_path = config.DB_JSON_PATH
    dir_name = file_path.parent
    dir_name.mkdir(parents=True, exist_ok=True)
    temp_name = None
    try:
        with tempfile.NamedTemporaryFile('w', dir=dir_name, delete=False, encoding='utf-8') as tf:
            temp_name = tf.name
            json.dump(history_db, tf, ensure_ascii=False, indent=2)
        os.replace(temp_name, file_path)
    except Exception as e:
        if temp_name and os.path.exists(temp_name):
            try:
                os.remove(temp_name)
            except Exception:
                pass
        raise e

def calculate_row_hash(r) -> str:
    # Combine all classification inputs to compute a row content hash
    # Ensure empty cells (NaN, None, "", "nan", "none") on both pandas and openpyxl result in the same empty string
    import pandas as pd
    parts = []
    for col in ["Tình trạng hiện tại", "Tình hình tiến độ công trình", 
                "Nội dung làm việc, yêu cầu KH & đánh giá", "Kế hoạch lần tới", "Đề xuất"]:
        val = r.get(col)
        if pd.isna(val) or val is None or str(val).strip().lower() in ("nan", "none", "null", ""):
            parts.append("")
        else:
            parts.append(str(val).strip())
    raw_str = "|".join(parts)
    return hashlib.md5(raw_str.encode("utf-8")).hexdigest()

def map_excel_columns(ws) -> dict:
    # Map column names to indices, handling merged headers in row 1
    col_mapping = {}
    for col_idx in range(1, ws.max_column + 1):
        major = str(ws.cell(row=1, column=col_idx).value or "").strip()
        minor = str(ws.cell(row=2, column=col_idx).value or "").strip()
        
        # Clean up suffixes like .1 from headers if any exist
        if "." in major:
            major = major.split(".")[0].strip()
        if "." in minor:
            minor = minor.split(".")[0].strip()
            
        # Merged cells check: If major is empty, scan left to find the merged header
        if not major and col_idx > 1:
            for left_idx in range(col_idx - 1, 0, -1):
                left_val = str(ws.cell(row=1, column=left_idx).value or "").strip()
                if left_val in ["Hoạt Động CRM", "AETT", "Khách Hàng", "Kế Hoạch", "Đối Thủ Cạnh Tranh"]:
                    major = left_val
                    break
        
        if major in ["Hoạt Động CRM", "AETT", "Khách Hàng", "Kế Hoạch", "Đối Thủ Cạnh Tranh"] and minor:
            col_name = f"[{major}] {minor}"
        else:
            col_name = major or minor
        if col_name:
            col_mapping[col_name] = col_idx
    return col_mapping

def _is_processed_marker(val) -> bool:
    return str(val or "").strip().lower() in ("1", "true", "yes", "processed")

def _ensure_cache_metadata_columns(ws, col_mapping: dict) -> dict:
    for col_name in CACHE_META_COLUMNS:
        col_idx = col_mapping.get(col_name)
        if not col_idx:
            col_idx = ws.max_column + 1
            ws.cell(row=1, column=col_idx, value=col_name)
            ws.cell(row=2, column=col_idx, value=col_name)
            col_mapping[col_name] = col_idx
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    return col_mapping

def _write_cache_metadata(ws, row_num: int, col_mapping: dict, fills: dict) -> None:
    processed_col = col_mapping.get(CACHE_META_PROCESSED)
    hash_col = col_mapping.get(CACHE_META_CONTENT_HASH)
    updated_col = col_mapping.get(CACHE_META_UPDATED_AT)

    if processed_col:
        ws.cell(row=row_num, column=processed_col, value="1")
    if hash_col:
        ws.cell(row=row_num, column=hash_col, value=fills.get("_content_hash"))
    if updated_col and not ws.cell(row=row_num, column=updated_col).value:
        ws.cell(
            row=row_num,
            column=updated_col,
            value=datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        )

def run_automation_pipeline() -> bool:
    logger.info("============================================================")
    logger.info("CRM Automated Pipeline Run Started")
    logger.info("============================================================")
    
    t0 = time.time()
    auth_provider = AuthProvider()
    sp_client = SharePointClient(auth_provider)
    notifier = NotificationService(auth_provider)
    
    local_excel_path = config.PATH_INPUT
    
    # Clean up local input directory if leftover
    if local_excel_path.exists():
        try:
            local_excel_path.unlink()
        except Exception:
            pass

    try:
        # 1. Download file from SharePoint Source site
        sp_client.download_file(
            config.SHAREPOINT_SOURCE_FILE_PATH, 
            local_excel_path, 
            drive_id=config.SHAREPOINT_SOURCE_DRIVE_ID
        )
        
        # 2. Schema Validation
        df = pd.read_excel(local_excel_path)
        required_cols = ["ActivityId", "Tình trạng hiện tại", "Tình hình tiến độ công trình", 
                         "Nội dung làm việc, yêu cầu KH & đánh giá", "Kế hoạch lần tới", "Đề xuất"]
        for col in required_cols:
            if col not in df.columns:
                raise ValueError(f"Schema mismatch: missing column '{col}'")

        # 3. Create Backup of the input file
        config.PATH_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = config.PATH_BACKUP_DIR / f"CRM_merge_backup_{timestamp}.xlsx"
        shutil.copy(local_excel_path, backup_path)
        logger.info("[OK] Backup created successfully: %s", backup_path.name)

        # 4. Load History DB (Priority 1: JSON)
        history_db = {}
        cache_loaded_from_json = False
        if config.DB_JSON_PATH.exists():
            try:
                with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
                    raw_db = json.load(f)
                # Normalize keys on the fly
                for k, v in raw_db.items():
                    k_norm = config.normalize_id(k)
                    if k_norm:
                        history_db[k_norm] = v
                logger.info("Loaded history DB from JSON: %d items", len(history_db))
                if history_db:
                    cache_loaded_from_json = True
            except Exception as e:
                logger.warning("Failed to load history DB JSON: %s.", e)

        # Priority 2: Rebuild cache from SharePoint target Excel file if JSON cache is empty/missing
        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        target_excel_path = config.PATH_OUTPUT / target_file_name
        target_downloaded_at_start = False

        if not cache_loaded_from_json:
            logger.info("JSON cache is empty or corrupt. Checking target file on SharePoint to rebuild cache...")
            target_exists = False
            try:
                target_exists = sp_client.check_file_exists(
                    config.SHAREPOINT_TARGET_FILE_PATH,
                    drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
                )
            except Exception as e:
                logger.warning("Failed to check target file existence on SharePoint: %s", e)

            if target_exists:
                logger.info("Target file exists on SharePoint. Downloading to rebuild JSON cache...")
                try:
                    if target_excel_path.exists():
                        target_excel_path.unlink()
                    sp_client.download_file(
                        config.SHAREPOINT_TARGET_FILE_PATH,
                        target_excel_path,
                        drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
                    )
                    target_downloaded_at_start = True
                    
                    # Rebuild history_db from the downloaded target Excel file
                    rebuilt_db = {}
                    try:
                        logger.info("Parsing target Excel file for cache reconstruction...")
                        wb = openpyxl.load_workbook(target_excel_path, data_only=True)
                        ws = wb.active
                        
                        # Map columns using the global helper
                        col_mapping = map_excel_columns(ws)
                                
                        act_id_col_idx = col_mapping.get("ActivityId")
                        if act_id_col_idx:
                            # Read input columns needed for hash calculation
                            input_cols = ["Tình trạng hiện tại", "Tình hình tiến độ công trình", 
                                          "Nội dung làm việc, yêu cầu KH & đánh giá", "Kế hoạch lần tới", "Đề xuất"]
                            
                            for r in range(3, ws.max_row + 1):
                                act_val = ws.cell(row=r, column=act_id_col_idx).value
                                act_id = config.normalize_id(act_val)
                                if not act_id:
                                    continue
                                    
                                # Check if classification columns already have non-empty, non-"mơ hồ" values
                                fills = {}
                                has_any_classification = False
                                processed_col_idx = col_mapping.get(CACHE_META_PROCESSED)
                                hash_col_idx = col_mapping.get(CACHE_META_CONTENT_HASH)
                                was_processed = (
                                    _is_processed_marker(ws.cell(row=r, column=processed_col_idx).value)
                                    if processed_col_idx
                                    else False
                                )
                                stored_hash = (
                                    str(ws.cell(row=r, column=hash_col_idx).value or "").strip()
                                    if hash_col_idx
                                    else ""
                                )
                                for col_name in config.OUTPUT_COLUMNS:
                                    col_idx = col_mapping.get(col_name)
                                    if col_idx:
                                        val = ws.cell(row=r, column=col_idx).value
                                        val_clean = str(val).strip() if val is not None else ""
                                        if val_clean != "" and val_clean.lower() != "mơ hồ":
                                            fills[col_name] = val_clean
                                            # Only count it as cached if at least one actual classification target column is filled
                                            if col_name != config.COL_CURRENT_STATUS:
                                                has_any_classification = True
                                
                                if has_any_classification or was_processed:
                                    # Read input values to calculate hash
                                    row_inputs = {}
                                    for col_name in input_cols:
                                        col_idx = col_mapping.get(col_name)
                                        if col_idx:
                                            row_inputs[col_name] = ws.cell(row=r, column=col_idx).value
                                    
                                    # Calculate hash using the global helper
                                    h_val = stored_hash or calculate_row_hash(row_inputs)
                                    
                                    rebuilt_db[act_id] = {
                                        **fills,
                                        "_content_hash": h_val
                                    }
                        wb.close()
                        logger.info("Reconstructed %d cached entries from Excel.", len(rebuilt_db))
                    except Exception as parse_err:
                        logger.error("Failed to parse target Excel for cache reconstruction: %s", parse_err)
                        
                    if rebuilt_db:
                        history_db.update(rebuilt_db)
                        # Save the reconstructed cache back as a healthy JSON file
                        save_history_db_atomic(history_db)
                        logger.info("Saved reconstructed cache to JSON history DB.")
                except Exception as dl_err:
                    logger.error("Failed to download or rebuild cache from target Excel: %s", dl_err)

        # 5. Extract Keywords Patterns & Filter Delta Rows
        kw_index, brand_list = load_and_index_keywords()
        col_pats = {col: {lbl: _compile_patterns(kws) for lbl, kws in labels_dict.items()}
                    for col, labels_dict in kw_index.items()}
        brand_pats = [(b, _compile_patterns([b])) for b in brand_list]

        # (Using global calculate_row_hash helper)

        pending_rows = []
        history_db_updated = False
        for idx, row in df.iterrows():
            act_id = config.normalize_id(row["ActivityId"])
            if not act_id:
                continue

            current_hash = calculate_row_hash(row)
            is_cached = False

            if act_id in history_db:
                cached_record = history_db[act_id]
                cached_hash = cached_record.get("_content_hash")
                if cached_hash == current_hash:
                    is_cached = True
                elif cached_hash is None:
                    # Legacy / seeded records: auto-upgrade hash to prevent unnecessary LLM runs
                    cached_record["_content_hash"] = current_hash
                    history_db_updated = True
                    is_cached = True
                    logger.info("Auto-upgraded legacy cache entry for ActivityId: %s", act_id)

            if not is_cached:
                pending_rows.append({
                    "ActivityId": act_id,
                    "row_idx": idx,
                    "Tình trạng hiện tại": row.get("Tình trạng hiện tại"),
                    "Tình hình tiến độ công trình": row.get("Tình hình tiến độ công trình"),
                    "Nội dung làm việc, yêu cầu KH & đánh giá": row.get("Nội dung làm việc, yêu cầu KH & đánh giá"),
                    "Kế hoạch lần tới": row.get("Kế hoạch lần tới"),
                    "Đề xuất": row.get("Đề xuất"),
                    "_content_hash": current_hash
                })

        if history_db_updated:
            logger.info("Saving history DB with auto-upgraded legacy content hashes...")
            try:
                save_history_db_atomic(history_db)
            except Exception as save_err:
                logger.warning("Failed to save history DB after auto-upgrade: %s", save_err)

        logger.info("Total rows in source file: %d | New delta rows to classify: %d", len(df), len(pending_rows))

        cells_filled_count = 0
        
        if pending_rows:
            # 5.1 Run Regex Classifier
            new_fills = {}
            llm_input_payload = []
            
            for item in pending_rows:
                act_id = item["ActivityId"]
                r_val = item.get("Tình trạng hiện tại")
                r_cleaned = _clean(r_val)
                row_fills = {
                    config.COL_CURRENT_STATUS: r_cleaned if r_cleaned else None,
                    "_content_hash": item["_content_hash"]
                }
                
                # Regex step
                regex_fills = classify_row_regex(item, col_pats, brand_pats)
                row_fills.update(regex_fills)
                
                # Check for LLM backup columns
                missing_cols = []
                for col in config.LLM_TARGET_COLS:
                    val = row_fills.get(col)
                    if val is None or val == "mơ hồ":
                        missing_cols.append(col)
                        
                if missing_cols:
                    # Check if the row has any input text at all to avoid wasting Gemini tokens
                    has_any_input = False
                    for col_name in ["Tình trạng hiện tại", "Tình hình tiến độ công trình", 
                                "Nội dung làm việc, yêu cầu KH & đánh giá", "Kế hoạch lần tới", "Đề xuất"]:
                        val = item.get(col_name)
                        if val is not None and not pd.isna(val) and str(val).strip() != "" and str(val).strip().lower() not in ("nan", "none", "null"):
                            has_any_input = True
                            break
                            
                    if has_any_input:
                        llm_input_payload.append({
                            "row_idx": act_id, # Use ActivityId
                            "Tình trạng hiện tại": item.get("Tình trạng hiện tại") or "",
                            "Tình hình tiến độ công trình": item.get("Tình hình tiến độ công trình") or "",
                            "Nội dung làm việc, yêu cầu KH & đánh giá": item.get("Nội dung làm việc, yêu cầu KH & đánh giá") or "",
                            "Kế hoạch lần tới": item.get("Kế hoạch lần tới") or "",
                            "Đề xuất": item.get("Đề xuất") or "",
                            "missing_cols": missing_cols
                        })
                    else:
                        # For blank rows, we map missing columns to None (empty)
                        for col in missing_cols:
                            row_fills[col] = None
                new_fills[act_id] = row_fills

            # 5.2 Call Gemini Client for empty/ambiguous columns
            if llm_input_payload:
                logger.info("Calling Gemini API for %d items...", len(llm_input_payload))
                client, model_name = init_llm_client()
                system_prompt = Path(config.PATH_PROMPT).read_text(encoding="utf-8")
                
                batch_size = config.BATCH_SIZE
                llm_results = {}
                total_items = len(llm_input_payload)
                
                batches = [llm_input_payload[i:i + batch_size] for i in range(0, total_items, batch_size)]
                total_batches = len(batches)
                
                workers = int(os.getenv("GEMINI_CONCURRENT_WORKERS") or "3")
                
                from concurrent.futures import ThreadPoolExecutor, as_completed
                import threading
                
                lock = threading.Lock()
                
                def worker_task(batch_idx, batch):
                    batch_num = batch_idx + 1
                    logger.info("Worker calling Gemini API for batch %d/%d (%d items)...", batch_num, total_batches, len(batch))
                    
                    input_ids = {str(item["row_idx"]) for item in batch}
                    success_res = []
                    
                    try:
                        # Try to call full batch with max_retry=3
                        res = call_llm_batch(client, model_name, system_prompt, batch, max_retry=3)
                        
                        # Filter successful items and identify missing/truncated ones
                        output_ids = set()
                        for item in res:
                            rid = str(item.get("row_idx") or item.get("idx") or "")
                            if rid in input_ids:
                                success_res.append(item)
                                output_ids.add(rid)
                        
                        missing_ids = input_ids - output_ids
                        if missing_ids:
                            logger.warning("Batch %d: %d rows missing/truncated from response. Retrying missing rows...", batch_num, len(missing_ids))
                            current_missing_items = [row for row in batch if str(row["row_idx"]) in missing_ids]
                            
                            # Retry missing items as a batch up to 3 times to save tokens
                            retry_attempts = 3
                            for attempt in range(1, retry_attempts + 1):
                                if not current_missing_items:
                                    break
                                logger.info("Batch %d: Retrying %d missing items as a batch (Attempt %d/%d)...", 
                                            batch_num, len(current_missing_items), attempt, retry_attempts)
                                try:
                                    retry_res = call_llm_batch(client, model_name, system_prompt, current_missing_items, max_retry=2)
                                    for item in retry_res:
                                        rid = str(item.get("row_idx") or item.get("idx") or "")
                                        if rid in missing_ids:
                                            success_res.append(item)
                                            output_ids.add(rid)
                                    
                                    # Recalculate remaining missing items
                                    still_missing_ids = {str(x["row_idx"]) for x in current_missing_items} - {str(x.get("row_idx") or x.get("idx") or "") for x in retry_res}
                                    current_missing_items = [row for row in current_missing_items if str(row["row_idx"]) in still_missing_ids]
                                except Exception as retry_err:
                                    logger.warning("Batch %d: Batch retry attempt %d failed: %s", batch_num, attempt, retry_err)
                            
                            if current_missing_items:
                                logger.error("Batch %d: %d items failed to classify after %d batch retries.", 
                                             batch_num, len(current_missing_items), retry_attempts)
                    except Exception as e:
                        logger.error("Batch %d: Complete batch failure after 3 retries: %s. Skipping items in this batch to save tokens.", batch_num, e)

                    with lock:
                        for item in success_res:
                            rid = str(item.get("row_idx") or item.get("idx") or "")
                            llm_fills = item.get("fills") or {}
                            llm_results[rid] = llm_fills
                            
                            # Merge back into new_fills in-place
                            if rid in new_fills:
                                fills = new_fills[rid]
                                for col in config.LLM_TARGET_COLS:
                                    if fills.get(col) == "mơ hồ" or fills.get(col) is None:
                                        val = llm_fills.get(col)
                                        if val is None or str(val).strip() == "":
                                            fills[col] = None
                                        else:
                                            fills[col] = str(val).strip()
                                history_db[rid] = fills
                                
                        # Save checkpoint incrementally
                        try:
                            save_history_db_atomic(history_db)
                        except Exception as save_err:
                            logger.warning("Failed to save history DB checkpoint: %s", save_err)
                    logger.info("Worker finished processing batch %d/%d.", batch_num, total_batches)
                            
                with ThreadPoolExecutor(max_workers=workers) as executor:
                    futures = [executor.submit(worker_task, idx, b) for idx, b in enumerate(batches)]
                    for future in as_completed(futures):
                        future.result()
                
                # Merge LLM fills back into regex records
                for act_id, fills in new_fills.items():
                    if act_id in llm_results:
                        llm_fills = llm_results[act_id]
                        for col in config.LLM_TARGET_COLS:
                            if fills.get(col) == "mơ hồ" or fills.get(col) is None:
                                val = llm_fills.get(col)
                                if val is None or str(val).strip() == "":
                                    fills[col] = None
                                else:
                                    fills[col] = str(val).strip()
            
                # Ensure no "mơ hồ" strings remain in the final classifications
                for act_id, fills in new_fills.items():
                    for col in config.LLM_TARGET_COLS:
                        if fills.get(col) == "mơ hồ":
                            fills[col] = None
            
            # Save new delta results to local JSON DB
            history_db.update(new_fills)
            save_history_db_atomic(history_db)
            logger.info("JSON History database successfully updated.")

        # 6. Tải file đích từ SharePoint Target site (hoặc khởi tạo từ df nếu chưa tồn tại)
        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        target_excel_path = config.PATH_OUTPUT / target_file_name
        
        if not target_downloaded_at_start:
            if target_excel_path.exists():
                try:
                    target_excel_path.unlink()
                except Exception:
                    pass

            target_exists = sp_client.check_file_exists(
                config.SHAREPOINT_TARGET_FILE_PATH, 
                drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
            )
            
            if not target_exists:
                logger.info("Target file does not exist on SharePoint. Initializing from source schema...")
                # If target file doesn't exist, we save source df to excel and style it as base template
                df.to_excel(target_excel_path, index=False)
                apply_excel_styling(target_excel_path)
                # Upload styled template to SharePoint
                sp_client.upload_file(
                    target_excel_path, 
                    config.SHAREPOINT_TARGET_FILE_PATH, 
                    drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
                )
            
            logger.info("Downloading target file from SharePoint Target site...")
            sp_client.download_file(
                config.SHAREPOINT_TARGET_FILE_PATH, 
                target_excel_path, 
                drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
            )
        
        # Open Excel workbook in openpyxl for in-place writing
        logger.info("Opening target Excel file for in-place update...")
        wb = openpyxl.load_workbook(target_excel_path)
        ws = wb.active
        
        # 1. Map column names to indices using the global helper
        col_mapping = map_excel_columns(ws)
        col_mapping = _ensure_cache_metadata_columns(ws, col_mapping)
                
        # 3. Map ActivityId to row numbers (data starts at row 3)
        act_id_col_idx = col_mapping.get("ActivityId")
        if not act_id_col_idx:
            raise ValueError("ActivityId column not found in target Excel sheet!")
            
        row_mapping = {}
        for r in range(3, ws.max_row + 1):
            act_id = config.normalize_id(ws.cell(row=r, column=act_id_col_idx).value)
            if act_id:
                row_mapping[act_id] = r
                
        # 4. Prepare source lookup for appending brand new rows
        source_df_clean = df.copy()
        source_df_clean["ActivityId_str"] = source_df_clean["ActivityId"].apply(config.normalize_id)
        source_lookup = source_df_clean.drop_duplicates(subset=["ActivityId_str"]).set_index("ActivityId_str")
        
        cells_filled_count = 0
        new_rows_count = 0
        
        # 5. In-place merge and append new rows
        logger.info("Performing in-place update and appending new rows...")
        for act_id, fills in history_db.items():
            if not act_id:
                continue
                
            if act_id in row_mapping:
                # Update existing row
                row_num = row_mapping[act_id]
                for col_name in config.OUTPUT_COLUMNS:
                    val = fills.get(col_name)
                    col_idx = col_mapping.get(col_name)
                    if col_idx:
                        cell = ws.cell(row=row_num, column=col_idx)
                        current_val = cell.value
                        
                        val_clean = str(val).strip() if val is not None else ""
                        curr_clean = str(current_val).strip().lower() if current_val is not None else ""
                        
                        # 1. If we have a valid classification value, write it to empty/mơ hồ cells
                        if val_clean != "" and val_clean.lower() != "mơ hồ":
                            if curr_clean in ("", "mơ hồ", "none", "nan"):
                                cell.value = val
                                cells_filled_count += 1
                        # 2. If classified as empty/null, explicitly clear any pre-existing "mơ hồ" placeholder
                        elif curr_clean == "mơ hồ":
                            cell.value = None
                            cells_filled_count += 1
                _write_cache_metadata(ws, row_num, col_mapping, fills)
            else:
                # Append brand new row at the end
                if act_id in source_lookup.index:
                    src_row = source_lookup.loc[act_id]
                    next_row = ws.max_row + 1
                    new_rows_count += 1
                    
                    # Write all columns for the new row and copy styles
                    for col_name, col_idx in col_mapping.items():
                        if col_name in config.OUTPUT_COLUMNS:
                            val = fills.get(col_name)
                        elif col_name in src_row.index:
                            val = src_row[col_name]
                        else:
                            val = None
                        
                        cell = ws.cell(row=next_row, column=col_idx, value=val)
                        
                        # Copy style from row 3 to preserve formatting
                        ref_cell = ws.cell(row=3, column=col_idx)
                        cell.font = copy(ref_cell.font)
                        cell.border = copy(ref_cell.border)
                        cell.alignment = copy(ref_cell.alignment)
                        cell.fill = copy(ref_cell.fill)

                    cells_filled_count += len(config.OUTPUT_COLUMNS)
                    _write_cache_metadata(ws, next_row, col_mapping, fills)
                    
        if new_rows_count:
            logger.info("Appended %d brand new rows to Excel sheet.", new_rows_count)
            
        # Save workbook locally
        wb.save(target_excel_path)
        logger.info("[OK] Excel updated locally in-place.")
        
        # 8. Upload file đích ngược lại SharePoint Target site
        sp_client.upload_file(
            target_excel_path, 
            config.SHAREPOINT_TARGET_FILE_PATH, 
            drive_id=config.SHAREPOINT_TARGET_DRIVE_ID
        )
        
        elapsed = time.time() - t0
        logger.info("[SUCCESS] Pipeline execution finished in %.1fs!", elapsed)
        
        # 9. Gửi email báo cáo thành công
        notifier.send_success(elapsed, len(pending_rows), cells_filled_count)
        
        # 10. Dọn dẹp các file Excel tạm cục bộ
        for p in (local_excel_path, target_excel_path):
            if p.exists():
                try:
                    p.unlink()
                except Exception:
                    pass
            
        return True

    except Exception as e:
        err_msg = traceback.format_exc()
        logger.error("Pipeline crashed! Exception traceback:\n%s", err_msg)
        
        # Gửi email báo cáo lỗi
        try:
            notifier.send_error(err_msg)
        except Exception as mail_err:
            logger.error("Failed to send error notification email: %s", mail_err)
            
        # Dọn dẹp tệp tạm
        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        for p in (local_excel_path, config.PATH_OUTPUT / target_file_name):
            if p.exists():
                try:
                    p.unlink()
                except Exception:
                    pass
                
        return False

def get_seconds_until_next_run(target_hour=3, target_minute=30) -> float:
    now = datetime.datetime.now()
    target_today = now.replace(hour=target_hour, minute=target_minute, second=0, microsecond=0)
    
    if now < target_today:
        diff = target_today - now
    else:
        target_tomorrow = target_today + datetime.timedelta(days=1)
        diff = target_tomorrow - now
        
    return diff.total_seconds()

if __name__ == "__main__":
    run_as_daemon = os.getenv("RUN_AS_DAEMON", "false").lower() == "true"
    
    if not run_as_daemon:
        success = run_automation_pipeline()
        sys.exit(0 if success else 1)
    else:
        logger.info("Starting pipeline in Daemon Mode (RUN_AS_DAEMON=True)...")
        # Run immediately on startup
        run_automation_pipeline()
        
        while True:
            sec = get_seconds_until_next_run(3, 30)
            hours = sec / 3600.0
            logger.info("Daemon Mode: Sleeping for %.2f hours until next run at 03:30 AM...", hours)
            time.sleep(sec)
            
            logger.info("Daemon Mode: Waking up to run scheduled pipeline...")
            run_automation_pipeline()
