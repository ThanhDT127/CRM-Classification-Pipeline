import sys
import os
import json
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch, mock_open
import pytest
import pandas as pd
import openpyxl

# Add src to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import config
from sharepoint import AuthProvider, SharePointClient
from notification import NotificationService
import pipeline as main

@pytest.fixture
def mock_auth():
    auth = MagicMock(spec=AuthProvider)
    auth.get_access_token.return_value = "mock_token"
    auth.get_headers.return_value = {"Authorization": "Bearer mock_token", "Content-Type": "application/json"}
    return auth

def test_sharepoint_client_download(mock_auth):
    client = SharePointClient(mock_auth)
    
    # Mock requests.Session
    client.session = MagicMock()
    
    # Mock metadata request
    mock_meta_resp = MagicMock()
    mock_meta_resp.status_code = 200
    mock_meta_resp.json.return_value = {
        "@microsoft.graph.downloadUrl": "https://mock.download.url/file.xlsx",
        "name": "CRM_merge.xlsx",
        "size": 1000
    }
    
    # Mock download request
    mock_dl_resp = MagicMock()
    mock_dl_resp.status_code = 200
    mock_dl_resp.iter_content.return_value = [b"mock excel content"]
    
    # Configure session.get side effects
    client.session.get.side_effect = [mock_meta_resp, mock_dl_resp]
    
    with tempfile.TemporaryDirectory() as tmpdir:
        local_path = Path(tmpdir) / "CRM_merge.xlsx"
        result_path = client.download_file("CRM_merge/CRM_merge.xlsx", local_path)
        
        assert result_path == local_path
        assert local_path.exists()
        assert local_path.read_bytes() == b"mock excel content"

def test_sharepoint_client_upload(mock_auth):
    client = SharePointClient(mock_auth)
    client.session = MagicMock()
    
    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.json.return_value = {"id": "mock_file_id", "name": "CRM_merge.xlsx"}
    client.session.put.return_value = mock_resp
    
    with tempfile.TemporaryDirectory() as tmpdir:
        local_path = Path(tmpdir) / "CRM_merge.xlsx"
        local_path.write_bytes(b"mock excel data to upload")
        
        res = client.upload_file(local_path, "CRM_merge/CRM_merge.xlsx")
        assert res["id"] == "mock_file_id"
        client.session.put.assert_called_once()

def test_notification_service_send_email(mock_auth):
    # Temporarily set config values
    config.NOTIFICATION_SENDER_EMAIL = "sender@domain.com"
    config.NOTIFICATION_RECIPIENTS = ["rec@domain.com"]
    
    notifier = NotificationService(mock_auth)
    notifier.session = MagicMock()
    
    mock_resp = MagicMock()
    mock_resp.status_code = 202
    notifier.session.post.return_value = mock_resp
    
    success = notifier._send_email("Test Subject", "<h1>Test Body</h1>")
    assert success is True
    notifier.session.post.assert_called_once()

@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_full_pipeline_run(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    print(">>> main module name:", main.__name__)
    print(">>> sys.modules pipeline keys:", [k for k in sys.modules if "pipeline" in k or "main" in k])
    mock_auth_inst = MagicMock()
    mock_auth_cls.return_value = mock_auth_inst
    
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    
    mock_notifier_inst = MagicMock()
    mock_notifier_cls.return_value = mock_notifier_inst
    
    mock_llm_client = MagicMock()
    mock_init_llm.return_value = (mock_llm_client, "gemini-2.5-flash")
    
    # Create temp files for config mapping
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        # Override config paths for testing
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"
        
        # Create a mock Excel sheet that needs classification
        df = pd.DataFrame([
            {
                "ActivityId": "ACT_001",
                "Tình trạng hiện tại": "Khách hàng muốn mua bóng đèn LED",
                "Tình hình tiến độ công trình": "Đang hoàn thiện phần thô",
                "Nội dung làm việc, yêu cầu KH & đánh giá": "Tư vấn giá và mẫu mã cho anh Thanh",
                "Kế hoạch lần tới": "Gửi báo giá bóng đèn",
                "Đề xuất": "Giảm giá 5%"
            }
        ])
        df.to_excel(config.PATH_INPUT, index=False)
        
        # Mock download_file to write the Excel file when called (since the pipeline unlinks it initially)
        def side_effect_download(remote_path, local_path, *args, **kwargs):
            df.to_excel(local_path, index=False)
            return local_path
        mock_sp_inst.download_file.side_effect = side_effect_download
        mock_sp_inst.check_file_exists.return_value = False
        
        # Mock LLM API response
        mock_call_llm.return_value = [
            {
                "row_idx": "ACT_001",
                "fills": {
                    config.COL_WORK_CRM: "Bóng LED",
                    config.COL_OPINION: "Muốn mua bóng đèn LED",
                    config.COL_PLAN_NEXT: "Gửi báo giá bóng",
                    config.COL_PROPOSAL: "Giảm giá 5%"
                }
            }
        ]
        
        # Run pipeline
        success = main.run_automation_pipeline()
        
        # Assertions
        assert success is True
        assert mock_sp_inst.download_file.call_count == 2
        assert mock_sp_inst.upload_file.call_count == 2
        mock_notifier_inst.send_success.assert_called_once()
        
        # Check that backup was created
        backups = list(config.PATH_BACKUP_DIR.glob("*.xlsx"))
        assert len(backups) == 1
        
        # Check history DB was updated
        assert config.DB_JSON_PATH.exists()
        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
            assert "ACT_001" in history
            assert history["ACT_001"][config.COL_CURRENT_STATUS] == "Khách hàng muốn mua bóng đèn LED"

@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_consecutive_pipeline_runs(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_inst = MagicMock()
    mock_auth_cls.return_value = mock_auth_inst
    
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    
    mock_notifier_inst = MagicMock()
    mock_notifier_cls.return_value = mock_notifier_inst
    
    mock_llm_client = MagicMock()
    mock_init_llm.return_value = (mock_llm_client, "gemini-2.5-flash")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        # Override config paths for testing
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"
        
        # Create a mock Excel sheet that needs classification
        df = pd.DataFrame([
            {
                "ActivityId": "ACT_001",
                "Tình trạng hiện tại": "Khách hàng muốn mua bóng đèn LED",
                "Tình hình tiến độ công trình": "Đang hoàn thiện phần thô",
                "Nội dung làm việc, yêu cầu KH & đánh giá": "Tư vấn giá và mẫu mã cho anh Thanh",
                "Kế hoạch lần tới": "Gửi báo giá bóng đèn",
                "Đề xuất": "Giảm giá 5%"
            }
        ])
        df.to_excel(config.PATH_INPUT, index=False)
        
        # Mock SharePoint storage
        mock_sharepoint_files = {}
        
        def side_effect_download(remote_path, local_path, *args, **kwargs):
            if remote_path in mock_sharepoint_files:
                with open(local_path, "wb") as f:
                    f.write(mock_sharepoint_files[remote_path])
            else:
                df.to_excel(local_path, index=False)
            return local_path
            
        def side_effect_upload(local_path, remote_path, *args, **kwargs):
            with open(local_path, "rb") as f:
                mock_sharepoint_files[remote_path] = f.read()
            return {"id": "mock_id"}
            
        def side_effect_exists(remote_path, *args, **kwargs):
            return remote_path in mock_sharepoint_files
            
        mock_sp_inst.download_file.side_effect = side_effect_download
        mock_sp_inst.upload_file.side_effect = side_effect_upload
        mock_sp_inst.check_file_exists.side_effect = side_effect_exists
        
        # Mock LLM API response
        mock_call_llm.return_value = [
            {
                "row_idx": "ACT_001",
                "fills": {
                    config.COL_WORK_CRM: "Bóng LED",
                    config.COL_OPINION: "Muốn mua bóng đèn LED",
                    config.COL_PLAN_NEXT: "Gửi báo giá bóng",
                    config.COL_PROPOSAL: "Giảm giá 5%"
                }
            }
        ]
        
        # --- RUN 1 (Classification + Append) ---
        success1 = main.run_automation_pipeline()
        assert success1 is True
        
        # Verify first run updated history and uploaded file
        assert config.DB_JSON_PATH.exists()
        assert config.SHAREPOINT_TARGET_FILE_PATH in mock_sharepoint_files
        
        # Reset mock call histories for Run 2
        mock_sp_inst.download_file.reset_mock()
        mock_sp_inst.upload_file.reset_mock()
        mock_notifier_inst.send_success.reset_mock()
        mock_call_llm.reset_mock()
        
        # --- RUN 2 (Delta = 0, direct in-place check, no LLM call) ---
        # Re-create input file since pipeline deletes it on finish
        df.to_excel(config.PATH_INPUT, index=False)
        
        success2 = main.run_automation_pipeline()
        assert success2 is True
        
        # Assertions for Run 2:
        # 1. download_file should be called 2 times (download source and download target from SharePoint)
        assert mock_sp_inst.download_file.call_count == 2
        # 2. upload_file should be called 1 time (upload final updated target file)
        assert mock_sp_inst.upload_file.call_count == 1
        # 3. Gemini LLM should NOT be called since delta is 0
        mock_call_llm.assert_not_called()


@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_legacy_cache_auto_upgrade(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_inst = MagicMock()
    mock_auth_cls.return_value = mock_auth_inst
    
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    
    mock_notifier_inst = MagicMock()
    mock_notifier_cls.return_value = mock_notifier_inst
    
    mock_llm_client = MagicMock()
    mock_init_llm.return_value = (mock_llm_client, "gemini-2.5-flash")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        # Override config paths for testing
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"
        
        # Create a mock Excel sheet that needs classification
        df = pd.DataFrame([
            {
                "ActivityId": "ACT_001",
                "Tình trạng hiện tại": "Khách hàng muốn mua bóng đèn LED",
                "Tình hình tiến độ công trình": "Đang hoàn thiện phần thô",
                "Nội dung làm việc, yêu cầu KH & đánh giá": "Tư vấn giá và mẫu mã cho anh Thanh",
                "Kế hoạch lần tới": "Gửi báo giá bóng đèn",
                "Đề xuất": "Giảm giá 5%"
            }
        ])
        df.to_excel(config.PATH_INPUT, index=False)
        
        # Create a legacy/seeded history DB record without _content_hash
        legacy_db = {
            "ACT_001": {
                config.COL_CURRENT_STATUS: "Khách hàng muốn mua bóng đèn LED",
                config.COL_WORK_CRM: "Bóng LED",
                config.COL_OPINION: "Muốn mua bóng đèn LED",
                config.COL_PLAN_NEXT: "Gửi báo giá bóng",
                config.COL_PROPOSAL: "Giảm giá 5%"
            }
        }
        with open(config.DB_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(legacy_db, f)
            
        # Mock SharePoint storage download
        def side_effect_download(remote_path, local_path, *args, **kwargs):
            df.to_excel(local_path, index=False)
            return local_path
        mock_sp_inst.download_file.side_effect = side_effect_download
        mock_sp_inst.check_file_exists.return_value = True
        
        # Run pipeline
        success = main.run_automation_pipeline()
        assert success is True
        
        # Gemini LLM should NOT be called since the legacy cache is auto-upgraded and hit
        mock_call_llm.assert_not_called()
        
        # Verify the database now contains the auto-upgraded _content_hash key
        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            updated_db = json.load(f)
            assert "ACT_001" in updated_db
            assert "_content_hash" in updated_db["ACT_001"]
            assert len(updated_db["ACT_001"]["_content_hash"]) == 32


@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_self_healing_fallback(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_inst = MagicMock()
    mock_auth_cls.return_value = mock_auth_inst
    
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    
    mock_notifier_inst = MagicMock()
    mock_notifier_cls.return_value = mock_notifier_inst
    
    mock_llm_client = MagicMock()
    mock_init_llm.return_value = (mock_llm_client, "gemini-2.5-flash")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        # Override config paths for testing
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"
        
        # Create a mock Excel sheet with 1 row that needs classification
        df = pd.DataFrame([
            {
                "ActivityId": "ACT_001",
                "Tình trạng hiện tại": "Khách hàng muốn mua bóng đèn LED",
                "Tình hình tiến độ công trình": "Đang hoàn thiện phần thô",
                "Nội dung làm việc, yêu cầu KH & đánh giá": "Tư vấn giá và mẫu mã cho anh Thanh",
                "Kế hoạch lần tới": "Gửi báo giá bóng đèn",
                "Đề xuất": "Giảm giá 5%"
            }
        ])
        df.to_excel(config.PATH_INPUT, index=False)
        
        # Mock SharePoint storage download
        def side_effect_download(remote_path, local_path, *args, **kwargs):
            df.to_excel(local_path, index=False)
            return local_path
        mock_sp_inst.download_file.side_effect = side_effect_download
        mock_sp_inst.check_file_exists.return_value = False
        
        # Configure side effect for call_llm_batch:
        # First call (full batch) throws exception (API error)
        # Second call (fallback for row ACT_001) succeeds
        mock_call_llm.side_effect = [
            RuntimeError("Gemini Batch API Error"), # full batch call fails
            [ # single row fallback call succeeds
                {
                    "row_idx": "ACT_001",
                    "fills": {
                        config.COL_WORK_CRM: "Bóng LED",
                        config.COL_OPINION: "Muốn mua bóng đèn LED",
                        config.COL_PLAN_NEXT: "Gửi báo giá bóng",
                        config.COL_PROPOSAL: "Giảm giá 5%"
                    }
                }
            ]
        ]
        
        # Run pipeline
        success = main.run_automation_pipeline()
        assert success is True
        
        # call_llm_batch should be called twice:
        # 1. Once for the batch (failing)
        # 2. Once for the individual row (succeeding in fallback)
        assert mock_call_llm.call_count == 2
        
        # Verify the database contains the correctly classified row
        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
            assert "ACT_001" in history
            assert history["ACT_001"][config.COL_WORK_CRM] == "Bóng LED"
            assert "_content_hash" in history["ACT_001"]


@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_rebuild_cache_ignores_current_status_only_rows(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_cls.return_value = MagicMock()
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    mock_notifier_cls.return_value = MagicMock()
    mock_init_llm.return_value = (MagicMock(), "gemini-2.5-flash")
    mock_call_llm.return_value = [
        {"row_idx": "ACT_001", "fills": {config.COL_WORK_CRM: "Tiếp cận ban đầu"}}
    ]

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"

        df = pd.DataFrame([{
            "ActivityId": "ACT_001",
            "Tình trạng hiện tại": "Thi công điện nước",
            "Tình hình tiến độ công trình": "Đang thi công ME",
            "Nội dung làm việc, yêu cầu KH & đánh giá": "Chờ tiến độ cấp đèn",
            "Kế hoạch lần tới": None,
            "Đề xuất": None,
        }])
        df.to_excel(config.PATH_INPUT, index=False)

        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        target_excel_path = config.PATH_OUTPUT / target_file_name

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ActivityId")
        ws.cell(row=2, column=1, value="ActivityId")
        ws.cell(row=1, column=2, value="Hoạt Động CRM")
        ws.cell(row=2, column=2, value="Tình hình hiện tại")
        ws.cell(row=1, column=3, value="AETT")
        ws.cell(row=2, column=3, value="Nội dung làm việc")
        ws.cell(row=1, column=4, value="Tình trạng hiện tại")
        ws.cell(row=1, column=5, value="Tình hình tiến độ công trình")
        ws.cell(row=1, column=6, value="Nội dung làm việc, yêu cầu KH & đánh giá")
        ws.cell(row=1, column=7, value="Kế hoạch lần tới")
        ws.cell(row=1, column=8, value="Đề xuất")
        ws.cell(row=3, column=1, value="ACT_001")
        ws.cell(row=3, column=2, value="Thi công điện nước")
        ws.cell(row=3, column=3, value="mơ hồ")
        ws.cell(row=3, column=4, value="Thi công điện nước")
        ws.cell(row=3, column=5, value="Đang thi công ME")
        ws.cell(row=3, column=6, value="Chờ tiến độ cấp đèn")
        wb.save(target_excel_path)

        mock_sp_inst.check_file_exists.return_value = True

        def side_effect_download(remote_path, local_path, *args, **kwargs):
            if "CRM_merge.xlsx" in str(remote_path):
                df.to_excel(local_path, index=False)
            else:
                wb.save(local_path)
            return local_path

        mock_sp_inst.download_file.side_effect = side_effect_download

        success = main.run_automation_pipeline()
        assert success is True
        assert mock_call_llm.called

        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
        assert history["ACT_001"][config.COL_WORK_CRM] == "Tiếp cận ban đầu"


@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_rebuild_cache_preserves_processed_null_rows_from_metadata(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_cls.return_value = MagicMock()
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    mock_notifier_cls.return_value = MagicMock()
    mock_init_llm.return_value = (MagicMock(), "gemini-2.5-flash")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"

        source_row = {
            "ActivityId": "ACT_001",
            "Tình trạng hiện tại": "Thi công điện nước",
            "Tình hình tiến độ công trình": "Đang thi công ME",
            "Nội dung làm việc, yêu cầu KH & đánh giá": "Chờ tiến độ cấp đèn",
            "Kế hoạch lần tới": None,
            "Đề xuất": None,
        }
        df = pd.DataFrame([source_row])
        df.to_excel(config.PATH_INPUT, index=False)
        content_hash = main.calculate_row_hash(source_row)

        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        target_excel_path = config.PATH_OUTPUT / target_file_name

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="ActivityId")
        ws.cell(row=2, column=1, value="ActivityId")
        ws.cell(row=1, column=2, value="Hoạt Động CRM")
        ws.cell(row=2, column=2, value="Tình hình hiện tại")
        ws.cell(row=1, column=3, value="AETT")
        ws.cell(row=2, column=3, value="Nội dung làm việc")
        ws.cell(row=1, column=4, value="Tình trạng hiện tại")
        ws.cell(row=1, column=5, value="Tình hình tiến độ công trình")
        ws.cell(row=1, column=6, value="Nội dung làm việc, yêu cầu KH & đánh giá")
        ws.cell(row=1, column=7, value="Kế hoạch lần tới")
        ws.cell(row=1, column=8, value="Đề xuất")
        ws.cell(row=1, column=9, value=main.CACHE_META_PROCESSED)
        ws.cell(row=2, column=9, value=main.CACHE_META_PROCESSED)
        ws.cell(row=1, column=10, value=main.CACHE_META_CONTENT_HASH)
        ws.cell(row=2, column=10, value=main.CACHE_META_CONTENT_HASH)
        ws.cell(row=3, column=1, value="ACT_001")
        ws.cell(row=3, column=2, value="Thi công điện nước")
        ws.cell(row=3, column=3, value=None)
        ws.cell(row=3, column=4, value="Thi công điện nước")
        ws.cell(row=3, column=5, value="Đang thi công ME")
        ws.cell(row=3, column=6, value="Chờ tiến độ cấp đèn")
        ws.cell(row=3, column=9, value="1")
        ws.cell(row=3, column=10, value=content_hash)
        wb.save(target_excel_path)

        mock_sp_inst.check_file_exists.return_value = True

        def side_effect_download(remote_path, local_path, *args, **kwargs):
            if "CRM_merge.xlsx" in str(remote_path):
                df.to_excel(local_path, index=False)
            else:
                wb.save(local_path)
            return local_path

        mock_sp_inst.download_file.side_effect = side_effect_download

        success = main.run_automation_pipeline()
        assert success is True
        mock_call_llm.assert_not_called()

        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
        assert "ACT_001" in history
        assert history["ACT_001"]["_content_hash"] == content_hash


@patch("pipeline.AuthProvider")
@patch("pipeline.SharePointClient")
@patch("pipeline.NotificationService")
@patch("pipeline.init_llm_client")
@patch("pipeline.call_llm_batch")
def test_rebuild_cache_from_excel(mock_call_llm, mock_init_llm, mock_notifier_cls, mock_sp_cls, mock_auth_cls):
    mock_auth_inst = MagicMock()
    mock_auth_cls.return_value = mock_auth_inst
    
    mock_sp_inst = MagicMock()
    mock_sp_cls.return_value = mock_sp_inst
    
    mock_notifier_inst = MagicMock()
    mock_notifier_cls.return_value = mock_notifier_inst
    
    mock_llm_client = MagicMock()
    mock_init_llm.return_value = (mock_llm_client, "gemini-2.5-flash")
    
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        
        # Override config paths for testing
        config.PATH_OUTPUT = tmp_path / "output"
        config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
        config.PATH_INPUT = tmp_path / "CRM_merge.xlsx"
        config.PATH_BACKUP_DIR = tmp_path / "backups"
        config.DB_JSON_PATH = config.PATH_OUTPUT / "classified_history_db.json"
        config.CKPT_JSON = config.PATH_OUTPUT / "llm_fills_checkpoint.json"
        
        # 1. Create a mock source Excel sheet
        df = pd.DataFrame([
            {
                "ActivityId": "ACT_001",
                "Tình trạng hiện tại": "Khách hàng muốn mua bóng đèn LED",
                "Tình hình tiến độ công trình": "Đang hoàn thiện phần thô",
                "Nội dung làm việc, yêu cầu KH & đánh giá": "Tư vấn giá và mẫu mã cho anh Thanh",
                "Kế hoạch lần tới": "Gửi báo giá bóng đèn",
                "Đề xuất": "Giảm giá 5%"
            }
        ])
        df.to_excel(config.PATH_INPUT, index=False)
        
        # 2. Create a mock target Excel sheet that already contains classifications
        target_file_name = Path(config.SHAREPOINT_TARGET_FILE_PATH).name
        target_excel_path = config.PATH_OUTPUT / target_file_name
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Row 1 headers (categories)
        ws.cell(row=1, column=1, value="ActivityId")
        ws.cell(row=1, column=2, value="AETT")
        ws.cell(row=1, column=3, value="Khách Hàng")
        ws.cell(row=1, column=4, value="Kế Hoạch")
        ws.cell(row=1, column=5, value="Kế Hoạch")
        ws.cell(row=1, column=6, value="Tình trạng hiện tại")
        ws.cell(row=1, column=7, value="Tình hình tiến độ công trình")
        ws.cell(row=1, column=8, value="Nội dung làm việc, yêu cầu KH & đánh giá")
        ws.cell(row=1, column=9, value="Kế hoạch lần tới")
        ws.cell(row=1, column=10, value="Đề xuất")
        
        # Row 2 headers (column names)
        ws.cell(row=2, column=1, value="ActivityId")
        ws.cell(row=2, column=2, value="Nội dung làm việc")
        ws.cell(row=2, column=3, value="Ý kiến KH")
        ws.cell(row=2, column=4, value="Kế hoạch lần tới")
        ws.cell(row=2, column=5, value="Đề xuất")
        ws.cell(row=2, column=6, value="Tình trạng hiện tại")
        ws.cell(row=2, column=7, value="Tình hình tiến độ công trình")
        ws.cell(row=2, column=8, value="Nội dung làm việc, yêu cầu KH & đánh giá")
        ws.cell(row=2, column=9, value="Kế hoạch lần tới")
        ws.cell(row=2, column=10, value="Đề xuất")
        
        # Row 3 (data)
        ws.cell(row=3, column=1, value="ACT_001")
        ws.cell(row=3, column=2, value="Bóng LED")
        ws.cell(row=3, column=3, value="Muốn mua bóng đèn LED")
        ws.cell(row=3, column=4, value="Gửi báo giá bóng")
        ws.cell(row=3, column=5, value="Giảm giá 5%")
        ws.cell(row=3, column=6, value="Khách hàng muốn mua bóng đèn LED")
        ws.cell(row=3, column=7, value="Đang hoàn thiện phần thô")
        ws.cell(row=3, column=8, value="Tư vấn giá và mẫu mã cho anh Thanh")
        ws.cell(row=3, column=9, value="Gửi báo giá bóng đèn")
        ws.cell(row=3, column=10, value="Giảm giá 5%")
        
        wb.save(target_excel_path)
        
        # Mock SharePoint storage: check_file_exists returns True
        mock_sp_inst.check_file_exists.return_value = True
        
        # Mock SharePoint storage download: copies df/target_excel to local destination
        def side_effect_download(remote_path, local_path, *args, **kwargs):
            if "CRM_merge.xlsx" in str(remote_path):
                df.to_excel(local_path, index=False)
            else:
                wb.save(local_path)
            return local_path
        mock_sp_inst.download_file.side_effect = side_effect_download
        
        # Ensure json cache does not exist initially
        assert not config.DB_JSON_PATH.exists()
        
        # Run pipeline
        success = main.run_automation_pipeline()
        assert success is True
        
        # Gemini LLM should NOT be called because the cache was rebuilt from target Excel and matched the source text
        mock_call_llm.assert_not_called()
        
        # Verify the database was created and contains the reconstructed cache
        assert config.DB_JSON_PATH.exists()
        with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
            history = json.load(f)
            assert "ACT_001" in history
            assert history["ACT_001"]["[AETT] Nội dung làm việc"] == "Bóng LED"
            assert history["ACT_001"]["[Khách Hàng] Ý kiến KH"] == "Muốn mua bóng đèn LED"
            assert "_content_hash" in history["ACT_001"]



