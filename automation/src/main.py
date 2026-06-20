import os
import json
import shutil
import datetime
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

import config
from step1_classify import (
    classify_row_regex,
    _compile_patterns,
    _clean
)
from step3_call_llm import init_llm_client, call_llm_batch

# Build keyword indices from configuration
def load_and_index_keywords():
    with open(config.PATH_KW_JSON, "r", encoding="utf-8") as f:
        kw = json.load(f)
        
    kw_index = {}
    competitor_brands = []
    
    for major, minors in (kw or {}).items():
        if not isinstance(minors, dict):
            continue
        for minor, subs in minors.items():
            if major == "Đối Thủ Cạnh Tranh" and minor == "Cạnh Tranh" and isinstance(subs, dict):
                brand_list = subs.get("Các Hãng cụ thể") or subs.get("Các Hãng đối thủ cạnh tranh")
                if isinstance(brand_list, list):
                    for b in brand_list:
                        for term in str(b).split(";"):
                            if term.strip():
                                competitor_brands.append(term.strip())
                continue
            if isinstance(subs, dict):
                col = config.col_name(major, minor)
                col_labels = {}
                for sub, kw_list in subs.items():
                    if not isinstance(sub, str) or not sub.strip():
                        continue
                    terms = []
                    if isinstance(kw_list, list):
                        for k in kw_list:
                            for t in str(k).split(";"):
                                if t.strip():
                                    terms.append(t.strip())
                    elif isinstance(kw_list, str):
                        terms = [t.strip() for t in kw_list.split(";") if t.strip()]
                    if terms:
                        col_labels[sub] = terms
                if col_labels:
                    kw_index[col] = col_labels
                    
    return kw_index, competitor_brands

def apply_excel_styling(file_path: Path):
    """Apply premium openpyxl styling matching user double-header requirements."""
    print(f"Applying premium two-row header styling to {file_path.name}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        group_colors = {
            "Hoạt Động CRM": "FFF2CC",
            "AETT": "E2EFDA",
            "Khách Hàng": "E8D5F5",
            "Kế Hoạch": "FCE4CC",
            "Đối Thủ Cạnh Tranh": "D6EAF8"
        }

        group_data_colors = {
            "Hoạt Động CRM": "FFFBEA",
            "AETT": "F2F9EE",
            "Khách Hàng": "F5EEFA",
            "Kế Hoạch": "FEF3EA",
            "Đối Thủ Cạnh Tranh": "EBF5FB"
        }

        font_major = Font(name="Segoe UI", size=11, bold=True)
        font_minor = Font(name="Segoe UI", size=10, bold=True)
        font_data = Font(name="Segoe UI", size=10)
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        ws.insert_rows(1, 1)
        
        col_info = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_val = str(ws.cell(row=2, column=col_idx).value or "")
            if cell_val.startswith("[") and "]" in cell_val:
                parts = cell_val.split("] ", 1)
                major = parts[0][1:]
                minor = parts[1]
                col_info[col_idx] = (major, minor)
                ws.cell(row=1, column=col_idx, value=major)
                ws.cell(row=2, column=col_idx, value=minor)
            else:
                col_info[col_idx] = (None, None)

        for col_idx in range(1, ws.max_column + 1):
            major, minor = col_info.get(col_idx, (None, None))
            
            if major:
                fill_color = group_colors.get(major, "FFFFFF")
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                
                c1 = ws.cell(row=1, column=col_idx)
                c1.fill = fill
                c1.font = font_major
                c1.alignment = align_center
                c1.border = thin_border
                
                c2 = ws.cell(row=2, column=col_idx)
                c2.fill = fill
                c2.font = font_minor
                c2.alignment = align_center
                c2.border = thin_border
            else:
                val = ws.cell(row=2, column=col_idx).value
                ws.cell(row=1, column=col_idx, value=val)
                ws.cell(row=2, column=col_idx, value="")
                
                fill_neutral = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                for r in (1, 2):
                    c = ws.cell(row=r, column=col_idx)
                    c.fill = fill_neutral
                    c.font = font_minor
                    c.alignment = align_center
                    c.border = thin_border
                    
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

            if major:
                data_color = group_data_colors.get(major, "FFFFFF")
                data_fill = PatternFill(start_color=data_color, end_color=data_color, fill_type="solid")
            else:
                data_fill = None

            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                cell.border = thin_border
                if data_fill and major:
                    cell.fill = data_fill

        start_col = None
        current_major = None
        for col_idx in range(1, ws.max_column + 2):
            major = col_info.get(col_idx, (None, None))[0]
            if major != current_major:
                if current_major is not None and start_col is not None:
                    end_col = col_idx - 1
                    if end_col > start_col:
                        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                if major is not None:
                    start_col = col_idx
                    current_major = major
                else:
                    start_col = None
                    current_major = None

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 24

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                val_str = str(ws.cell(row=row_idx, column=col_idx).value or "")
                lines = val_str.split("\n")
                for line in lines:
                    if len(line) > max_len:
                        max_len = len(line)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(file_path)
        print(f"Styling successfully applied to {file_path.name}!")
    except Exception as e:
        print(f"Error styling Excel: {e}")

def run_pipeline():
    print("=" * 60)
    print("CRM Automated Classification Pipeline - Run Start")
    print(f"Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. Schema Check
    if not config.PATH_INPUT.exists():
        print(f"[CRITICAL] Input Excel file not found at: {config.PATH_INPUT}")
        return

    df = pd.read_excel(config.PATH_INPUT)
    required_cols = ["ActivityId", "Tình trạng hiện tại", "Tình hình tiến độ công trình", 
                     "Nội dung làm việc, yêu cầu KH & đánh giá", "Kế hoạch lần tới", "Đề xuất"]
    for col in required_cols:
        if col not in df.columns:
            print(f"[CRITICAL ERROR] Sheet columns layout mismatch! Missing column: '{col}'. Stopping run.")
            return

    # 2. Automated backup
    config.PATH_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = config.PATH_BACKUP_DIR / f"CRM_merge_backup_{timestamp}.xlsx"
    shutil.copy(config.PATH_INPUT, backup_path)
    print(f"[OK] Created backup of source file at: {backup_path.name}")

    # 3. Load long-term history database
    config.PATH_OUTPUT.mkdir(parents=True, exist_ok=True)
    history_db = {}
    if config.DB_JSON_PATH.exists():
        try:
            with open(config.DB_JSON_PATH, "r", encoding="utf-8") as f:
                history_db = json.load(f)
            print(f"[OK] Loaded history DB: {len(history_db)} unique activities.")
        except Exception as e:
            print(f"[WARN] Error reading history DB: {e}. Starting fresh.")

    # 4. Filter new rows (delta check)
    kw_index, brand_list = load_and_index_keywords()
    col_pats = {col: {lbl: _compile_patterns(kws) for lbl, kws in labels_dict.items()}
                for col, labels_dict in kw_index.items()}
    brand_pats = [(b, _compile_patterns([b])) for b in brand_list]

    pending_rows = []
    for idx, row in df.iterrows():
        act_id = str(row["ActivityId"]).strip()
        if not act_id or act_id.lower() in ("nan", "none", ""):
            continue
        
        # Incremental check: only process if ID is not in history
        if act_id not in history_db:
            pending_rows.append({
                "ActivityId": act_id,
                "row_idx": idx,
                "Tình trạng hiện tại": row.get("Tình trạng hiện tại"),
                "Tình hình tiến độ công trình": row.get("Tình hình tiến độ công trình"),
                "Nội dung làm việc, yêu cầu KH & đánh giá": row.get("Nội dung làm việc, yêu cầu KH & đánh giá"),
                "Kế hoạch lần tới": row.get("Kế hoạch lần tới"),
                "Đề xuất": row.get("Đề xuất")
            })

    print(f"Total spreadsheet rows: {len(df)}")
    print(f"Incremental new rows to process: {len(pending_rows)}")

    if not pending_rows:
        print("[OK] No new rows to classify today. Writing final styled sheet.")
    else:
        # 5. Process new rows
        # Step 5.1: Run regex classifier
        new_fills = {}
        llm_input_payload = []
        
        for item in pending_rows:
            act_id = item["ActivityId"]
            # Copy Tình trạng hiện tại directly
            r_val = item.get("Tình trạng hiện tại")
            r_cleaned = _clean(r_val)
            
            row_fills = {config.COL_CURRENT_STATUS: r_cleaned if r_cleaned else None}
            
            # Apply regex classifier
            regex_fills = classify_row_regex(item, col_pats, brand_pats)
            row_fills.update(regex_fills)
            
            # Find which columns are missing or "mơ hồ" for LLM fallback
            missing_cols = []
            for col in config.LLM_TARGET_COLS:
                val = row_fills.get(col)
                if val is None or val == "mơ hồ":
                    missing_cols.append(col)
                    
            if missing_cols:
                llm_input_payload.append({
                    "row_idx": act_id, # Use ActivityId as the row identifier for LLM calls
                    "Tình trạng hiện tại": item.get("Tình trạng hiện tại") or "",
                    "Tình hình tiến độ công trình": item.get("Tình hình tiến độ công trình") or "",
                    "Nội dung làm việc, yêu cầu KH & đánh giá": item.get("Nội dung làm việc, yêu cầu KH & đánh giá") or "",
                    "Kế hoạch lần tới": item.get("Kế hoạch lần tới") or "",
                    "Đề xuất": item.get("Đề xuất") or "",
                    "missing_cols": missing_cols
                })
                
            new_fills[act_id] = row_fills

        # Step 5.2: Call Vertex AI / Gemini API for remaining columns
        if llm_input_payload:
            print(f"Calling Gemini for {len(llm_input_payload)} items...")
            client, model_name = init_llm_client()
            system_prompt = Path(config.PATH_PROMPT).read_text(encoding="utf-8")
            
            # Run in batches
            batch_size = config.BATCH_SIZE
            llm_results = {}
            for i in range(0, len(llm_input_payload), batch_size):
                batch = llm_input_payload[i:i + batch_size]
                try:
                    res = call_llm_batch(client, model_name, system_prompt, batch)
                    for item in res:
                        rid = str(item.get("row_idx"))
                        llm_results[rid] = item.get("fills") or {}
                except Exception as e:
                    print(f"[ERROR] LLM batch {i} failed: {e}")
            
            # Merge LLM fills back into delta records
            for act_id, fills in new_fills.items():
                if act_id in llm_results:
                    llm_fills = llm_results[act_id]
                    for col, val in llm_fills.items():
                        if col in config.LLM_TARGET_COLS:
                            # If LLM returns null/None, clean up and overwrite any "mơ hồ" tag
                            if val is None or str(val).strip() == "":
                                fills[col] = None
                            else:
                                fills[col] = str(val).strip()
            
        # Update history database
        history_db.update(new_fills)
        with open(config.DB_JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(history_db, f, ensure_ascii=False, indent=2)
        print(f"[OK] History DB updated with {len(new_fills)} new rows.")

    # 6. Gộp kết quả an toàn bằng Index (Merge safely by ActivityId index)
    for col in config.OUTPUT_COLUMNS:
        if col not in df.columns:
            df[col] = None
        df[col] = df[col].astype(object)

    df.set_index("ActivityId", inplace=True)
    
    merge_count = 0
    for act_id, fills in history_db.items():
        if act_id in df.index:
            for col in config.OUTPUT_COLUMNS:
                val = fills.get(col)
                current_val = df.at[act_id, col]
                is_empty_or_mo_ho = pd.isna(current_val) or str(current_val).strip() == '' or str(current_val).strip().lower() == 'mơ hồ'
                
                if is_empty_or_mo_ho and val is not None:
                    df.at[act_id, col] = val
                    merge_count += 1
                elif is_empty_or_mo_ho and val is None:
                    # Clear "mơ hồ" placeholder to empty cell
                    df.at[act_id, col] = None

    df.reset_index(inplace=True)

    # 7. Write and Style Excel
    df.to_excel(config.PATH_INPUT, index=False)
    print(f"[OK] Merged {merge_count} fields safely into Excel.")
    
    apply_excel_styling(config.PATH_INPUT)
    print("=" * 60)
    print("[SUCCESS] Pipeline Daily Run Completed!")
    print("=" * 60)

if __name__ == "__main__":
    run_pipeline()
